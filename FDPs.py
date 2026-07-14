"""
Ajuste de FDPs para el TP de la planta de soldadura, usando la biblioteca `fitter`.

Requiere: pip install fitter pandas openpyxl scipy matplotlib

Uso:
    python FDPs.py /ruta/al/reporte_produccion.xlsx
    (por defecto, se usa "Dataset.xlsx" en el mismo directorio)
"""
import sys
import numpy as np
import openpyxl
from datetime import datetime
from fitter import Fitter

ARCHIVO = sys.argv[1] if len(sys.argv) > 1 else "Dataset.xlsx"

# --------------------------------------------------------------------------- #
# 1. Carga de datos reales (mismo criterio que la propuesta: excluye
#    PEDESTAL 18 que está vacío)
# --------------------------------------------------------------------------- #
def cargar_filas(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    pedestales = [s for s in wb.sheetnames if s.startswith("PEDESTAL") and s != "PEDESTAL 18"]
    filas = []
    for sn in pedestales:
        ws = wb[sn]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0] is None or r[6] != "PROD":
                continue
            filas.append({
                "pedestal": sn,
                "fecha": r[0],
                "hora_ini": r[3],
                "dur": r[5],
                "pz1": r[8] or 0,
                "pzr": r[9] or 0,
            })
    return filas


def calcular_iap_tprod(filas):
    times = [datetime.strptime(r["hora_ini"], "%Y-%m-%d %H:%M:%S") for r in filas]
    times.sort()
    iap = [(times[i + 1] - times[i]).total_seconds() / 60 for i in range(len(times) - 1)]
    iap = [x for x in iap if 0 < x < 60]  # recortamos outliers de fin de semana

    tprod = []
    for r in filas:
        h, m, s = str(r["dur"]).split(":")
        tprod.append((int(h) * 3600 + int(m) * 60 + int(s)) / 60)
    tprod = [x for x in tprod if 0 < x < 120]

    return np.array(iap), np.array(tprod)


def tasa_rechazo(filas):
    pz1 = sum(r["pz1"] for r in filas)
    pzr = sum(r["pzr"] for r in filas)
    return pzr / pz1 if pz1 else 0.0


DISTRIBUCIONES = [
    "expon",
    "gamma",
    "lognorm",
    "weibull_min",
    "beta",
    "chi2",
    "triang",
    "wrapcauchy"
]


def ajustar(datos, nombre):
    print(f"\n=== Ajustando FDP de {nombre} (n={len(datos)}) ===")
    f = Fitter(datos, distributions=DISTRIBUCIONES)
    f.fit()
    print(f.summary(5, plot=False))
    mejor = f.get_best(method="sumsquare_error")
    print(f"MEJOR AJUSTE {nombre}: {mejor}")
    return mejor


if __name__ == "__main__":
    filas = cargar_filas(ARCHIVO)
    iap, tprod = calcular_iap_tprod(filas)
    tr = tasa_rechazo(filas)

    mejor_iap = ajustar(iap, "IAP (Intervalo de Arribo de Pedidos)")
    mejor_tprod = ajustar(tprod, "TPROD (Tiempo de Producción por pedido)")

    # --- Ajuste para PZ1 (cantidad PZ1 por pedido) -------------------------
    pz1_vals = np.array([r["pz1"] for r in filas], dtype=float)
    # Filtrar valores no positivos si la distribución debe modelar cantidades >0
    if len(pz1_vals) == 0:
        print("No hay datos de PZ1 para ajustar.")
    else:
        print(f"\nPZ1: n={len(pz1_vals)}, media={np.mean(pz1_vals):.3f}, mediana={np.median(pz1_vals):.3f}")
        mejor_pz1 = ajustar(pz1_vals, "PZ1 (cantidad por pedido)")

    print(f"\nTasa de rechazo real (PR): {tr*100:.3f}%")
    print(f"Pedestales productivos activos en el dataset: {len(set(r['pedestal'] for r in filas))}")
